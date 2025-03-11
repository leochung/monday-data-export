import json
import requests
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import csv
import openpyxl
from openpyxl.styles import Font

def read_config(config_file):
    """Reads configuration from a JSON file."""
    with open(config_file, 'r') as file:
        return json.load(file)

def read_api_token(file_path):
    """Reads API token from a file."""
    with open(file_path, 'r') as file:
        return file.read().strip()

def fetch_monday_board(api_token, board_name):
    """Fetches data from Monday.com board."""
    headers = {
        'Authorization': api_token,
        'Content-Type': 'application/json',
    }
    
    # First, get the board ID
    query = {
        "query": """
        {
            boards {
                id
                name
            }
        }
        """
    }
    
    response = requests.post('https://api.monday.com/v2', json=query, headers=headers)
    response_data = response.json()
    
    if 'errors' in response_data:
        print("Monday.com API Error:")
        for error in response_data['errors']:
            print(f"- {error.get('message')}")
            if 'locations' in error:
                for loc in error['locations']:
                    print(f"  at line {loc.get('line')}, column {loc.get('column')}")
        return None
    
    board_id = None
    for board in response_data['data']['boards']:
        if board['name'] == board_name:
            board_id = board['id']
            break
    
    if not board_id:
        print(f"Board '{board_name}' not found")
        return None
    
    # Now fetch all items using pagination
    all_items = []
    cursor = None
    
    # First page
    query = {
        "query": f"""
        {{
            boards(ids: {board_id}) {{
                name
                columns {{
                    title
                    id
                    type
                }}
                items_page(limit: 100) {{
                    cursor
                    items {{
                        id
                        name
                        column_values {{
                            id
                            type
                            text
                            column {{
                                title
                            }}
                        }}
                    }}
                }}
            }}
        }}
        """
    }
    
    response = requests.post('https://api.monday.com/v2', json=query, headers=headers)
    response_data = response.json()
    
    if 'errors' in response_data:
        print("Monday.com API Error:")
        for error in response_data['errors']:
            print(f"- {error.get('message')}")
            if 'locations' in error:
                for loc in error['locations']:
                    print(f"  at line {loc.get('line')}, column {loc.get('column')}")
        return None
    
    board_data = response_data['data']['boards'][0]
    items_page = board_data['items_page']
    all_items.extend(items_page['items'])
    cursor = items_page['cursor']
    
    # Subsequent pages
    while cursor:
        query = {
            "query": f"""
            {{
                next_items_page(cursor: "{cursor}", limit: 100) {{
                    cursor
                    items {{
                        id
                        name
                        column_values {{
                            id
                            type
                            text
                            column {{
                                title
                            }}
                        }}
                    }}
                }}
            }}
            """
        }
        
        response = requests.post('https://api.monday.com/v2', json=query, headers=headers)
        response_data = response.json()
        
        if 'errors' in response_data:
            print("Monday.com API Error:")
            for error in response_data['errors']:
                print(f"- {error.get('message')}")
                if 'locations' in error:
                    for loc in error['locations']:
                        print(f"  at line {loc.get('line')}, column {loc.get('column')}")
            break
        
        items_page = response_data['data']['next_items_page']
        all_items.extend(items_page['items'])
        cursor = items_page['cursor']
    
    # Return the data with all items
    return {
        'data': {
            'boards': [{
                'items': all_items,
                'columns': board_data['columns']
            }]
        }
    }

def write_to_csv(filename, data, table_header_id, exclusion_list):
    """Writes data to a CSV file with row exclusion and header deduplication."""
    with open(filename, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        
        # Extract headers from the first item
        headers = [table_header_id, "Item Name"] + [col['column']['title'] for col in data[0]['column_values']]
        writer.writerow(headers)
        
        # Process and write data rows
        seen_headers = set([tuple(headers)])  # Keep track of header rows we've seen
        for item in data:
            # Create the row data
            row = [item['id'], item['name']] + [col['text'] for col in item['column_values']]
            
            # Skip if this is a duplicate header row
            if tuple(row) in seen_headers:
                continue
            
            # Skip if the first cell (item id) matches any string in the exclusion list
            if any(row[0] == excl for excl in exclusion_list):
                continue
            
            # Add row to seen headers if it matches the header pattern
            if row[0] == table_header_id:
                seen_headers.add(tuple(row))
                continue
            
            # Write the row
            writer.writerow(row)

def write_to_excel(filename, data, table_header_id, exclusion_list):
    """Writes data to an Excel file with row exclusion and header deduplication.
    
    Args:
        filename: Name of the Excel file to write to.
        data: List of items from Monday.com board.
        table_header_id: String that identifies the table header row.
        exclusion_list: List of strings to match for row exclusion.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Monday Data"
    
    # Extract headers
    headers = [table_header_id, "Item Name"] + [col['column']['title'] for col in data[0]['column_values']]
    
    # Write headers with bold font
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
    
    # Process and write data rows
    seen_headers = set([tuple(headers)])  # Keep track of header rows we've seen
    row_num = 2
    
    for item in data:
        # Create the row data
        row = [item['id'], item['name']] + [col['text'] for col in item['column_values']]
        
        # Skip if this is a duplicate header row
        if tuple(row) in seen_headers:
            continue
        
        # Skip if the first cell matches any string in the exclusion list
        if any(row[0] == excl for excl in exclusion_list):
            continue
        
        # Add row to seen headers if it matches the header pattern
        if row[0] == table_header_id:
            seen_headers.add(tuple(row))
            continue
        
        # Write the row
        for col, value in enumerate(row, 1):
            ws.cell(row=row_num, column=col, value=value)
        row_num += 1
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save the workbook
    wb.save(filename)

def write_to_google_sheets(credentials_file, sheet_name, data, table_header_id, exclusion_list):
    """Writes data to Google Sheets with row exclusion and header deduplication."""
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    creds = ServiceAccountCredentials.from_json_keyfile_name(credentials_file, scope)
    client = gspread.authorize(creds)
    sheet = client.open(sheet_name).sheet1
    
    # Clear existing content
    sheet.clear()
    
    # Extract headers
    headers = [table_header_id, "Item Name"] + [col['column']['title'] for col in data[0]['column_values']]
    sheet.append_row(headers)
    
    # Process and write data rows
    seen_headers = set([tuple(headers)])  # Keep track of header rows we've seen
    for item in data:
        # Create the row data
        row = [item['id'], item['name']] + [col['text'] for col in item['column_values']]
        
        # Skip if this is a duplicate header row
        if tuple(row) in seen_headers:
            continue
        
        # Skip if the first cell (item id) matches any string in the exclusion list
        if any(row[0] == excl for excl in exclusion_list):
            continue
        
        # Add row to seen headers if it matches the header pattern
        if row[0] == table_header_id:
            seen_headers.add(tuple(row))
            continue
        
        # Write the row
        sheet.append_row(row)

if __name__ == "__main__":
    config = read_config("monday-data-export.conf")
    
    monday_api_token = read_api_token(config["api_token_path"])
    board_name = config["board_name"]
    
    monday_data = fetch_monday_board(monday_api_token, board_name)
    if monday_data and 'data' in monday_data:
        items = monday_data['data']['boards'][0]['items']
        
        output_format = config.get("output_format", "google_sheets")  # Default to google_sheets if not specified
        table_header_id = config.get("table_header_id", "Item ID")  # Default to "Item ID" if not specified
        exclusion_list = config.get("exclusion_list", [])  # Default to empty list if not specified
        
        if output_format == "google_sheets":
            google_credentials = config["google_credentials"]
            sheet_name = config["google_sheet_name"]
            write_to_google_sheets(google_credentials, sheet_name, items, table_header_id, exclusion_list)
            print("Data successfully written to Google Sheets!")
        elif output_format == "csv":
            filename = "monday_data.csv"
            write_to_csv(filename, items, table_header_id, exclusion_list)
            print(f"Data successfully written to {filename}!")
        elif output_format == "excel":
            filename = "monday_data.xlsx"
            write_to_excel(filename, items, table_header_id, exclusion_list)
            print(f"Data successfully written to {filename}!")
        else:
            print("Invalid output format specified in config. Please choose 'google_sheets', 'csv', or 'excel'.")
    else:
        print("Failed to fetch data from Monday.com")
